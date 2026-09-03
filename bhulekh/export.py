"""Excel / CSV exports."""
from __future__ import annotations

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .store import Store, split_label

HIT_COLS = ["category", "target", "score", "district", "tehsil", "village", "village_code", "khata", "khatedar",
            "father", "area_ha", "unique_code", "fasli", "name_score", "father_score", "reasoning", "pdf", "png"]


def _hit_row(h) -> list:
    return [h["category"], h["target"], round(h["score"], 1), split_label(h["district"])[0], split_label(h["tehsil"])[0],
            h["village_label"], h["village_code"], h["khata"], h["khatedar"], h["father"], h["area"], h["unique_code"],
            "current" if h["fasli"] == "999" else h["fasli"], round(h["name_score"], 1), round(h["father_score"], 1),
            h["reasoning"], h["pdf_path"] or "", h["png_path"] or ""]


def _autosize(ws):
    for col in ws.columns:
        width = min(60, max(10, max(len(str(c.value or "")) for c in col) + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def export_hits(store: Store, out_dir: str) -> str:
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "hits"
    ws.append(HIT_COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
    fill = {"probable": PatternFill("solid", fgColor="C6EFCE"), "less_probable": PatternFill("solid", fgColor="FFEB9C")}
    for h in store.hits():
        ws.append(_hit_row(h))
        ws.cell(row=ws.max_row, column=1).fill = fill.get(h["category"], PatternFill())
    _autosize(ws)
    cov = wb.create_sheet("coverage")
    cov.append(["district", "villages", "done", "errors", "skipped", "pending", "percent"])
    for r in store.coverage():
        cov.append([split_label(r["district"])[0], r["total"], r["done"], r["errors"], r["skipped"], r["pending"],
                    round(100.0 * ((r["done"] or 0) + (r["skipped"] or 0)) / max(r["total"], 1), 1)])
    _autosize(cov)
    path = os.path.join(out_dir, "hits.xlsx")
    wb.save(path)
    csv_path = os.path.join(out_dir, "hits.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HIT_COLS)
        for h in store.hits():
            w.writerow(_hit_row(h))
    return path


def export_district(store: Store, district_label: str, out_dir: str) -> str:
    os.makedirs(out_dir, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "rows"
    ws.append(["tehsil", "village", "village_code", "prefix", "fasli", "khata", "khatedar", "father", "area_ha", "unique_code"])
    for c in ws[1]:
        c.font = Font(bold=True)
    q = ("SELECT v.tehsil, v.label, r.* FROM rows r JOIN villages v ON v.code=r.village_code "
         "WHERE v.district=? ORDER BY v.tehsil, v.label, r.khatedar")
    for r in store.conn.execute(q, (district_label,)):
        ws.append([split_label(r["tehsil"])[0], r["label"], r["village_code"], r["prefix"],
                   "current" if r["fasli"] == "999" else r["fasli"], r["khata"], r["khatedar"], r["father"],
                   r["area"], r["unique_code"]])
    _autosize(ws)
    hs = wb.create_sheet("hits")
    hs.append(HIT_COLS)
    for h in store.hits():
        if h["district"] == district_label:
            hs.append(_hit_row(h))
    _autosize(hs)
    name = split_label(district_label)[0].replace(" ", "_")
    path = os.path.join(out_dir, f"{name}.xlsx")
    wb.save(path)
    return path
